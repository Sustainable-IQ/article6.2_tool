import json, csv, datetime, re
from openpyxl import load_workbook

SRC = "/home/claude/article6/src/Article-6.2-CA-Reporting-Tool_v1.xlsx"
OUT = "/home/claude/article6/data"

wbv = load_workbook(SRC, data_only=True)   # cached values
wbf = load_workbook(SRC, data_only=False)  # formulas

def iso(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return v

# ---------- 1. Detailed Credit Data ----------
sv = wbv["Detailed Credit Data"]
sf = wbf["Detailed Credit Data"]
headers = [c.value for c in sv[1]][:15]
rows = []
for i in range(2, sv.max_row + 1):
    vals = [iso(sv.cell(i, j).value) for j in range(1, 16)]
    if all(v is None or v == "" for v in vals):
        continue
    rec = dict(zip(headers, vals))
    # hyperlink target from the formula workbook
    hl = sf.cell(i, 15).hyperlink
    rec["Authorisation URL"] = hl.target if hl else None
    rows.append(rec)

with open(f"{OUT}/credits.json", "w") as f:
    json.dump(rows, f, indent=1, ensure_ascii=False)

cols = headers + ["Authorisation URL"]
with open(f"{OUT}/credits.csv", "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=cols)
    w.writeheader()
    w.writerows(rows)

# ---------- 2. Structured Summary meta ----------
ssv, ssf = wbv["Structured Summary"], wbf["Structured Summary"]
ss_rows = []
for i in range(3, 15):
    ss_rows.append({
        "excel_row": i,
        "table4_field": ssv.cell(i, 1).value,
        "explanation": ssv.cell(i, 8).value,
        "formula_col_B": ssf.cell(i, 2).value if isinstance(ssf.cell(i, 2).value, str) and str(ssf.cell(i, 2).value).startswith("=") else None,
        "cached_values": {str(ssv.cell(2, j).value): ssv.cell(i, j).value for j in range(2, 8)},
    })

# ---------- 3. Introduction ----------
iv, ifm = wbv["Introduction"], wbf["Introduction"]
intro = {
    "cover_text": iv["A1"].value,
    "how_to_use": iv["A2"].value,
    "questions": [{"cell": f"B{r}", "label": iv.cell(r, 1).value, "shipped_answer": iv.cell(r, 2).value} for r in range(5, 10)],
}

# ---------- 4. Validations ----------
validations = {}
for name in wbf.sheetnames:
    ws = wbf[name]
    dvs = []
    for dv in ws.data_validations.dataValidation:
        dvs.append({"ranges": str(dv.sqref), "type": dv.type, "formula1": dv.formula1})
    if dvs:
        validations[name] = dvs

# ---------- 5. Notes + Change Log ----------
nv = wbv["Notes"]
notes = [{"country": nv.cell(r, 1).value, "note": nv.cell(r, 2).value, "sources": nv.cell(r, 3).value}
         for r in range(2, nv.max_row + 1) if nv.cell(r, 1).value]
cl = wbv["Change Log"]
changelog = [{"version": cl.cell(r, 1).value, "date": iso(cl.cell(r, 2).value), "description": cl.cell(r, 3).value}
             for r in range(2, cl.max_row + 1) if cl.cell(r, 1).value]

meta = {
    "source_file": "Article-6.2-CA-Reporting-Tool_v1.xlsx",
    "publishers": ["Gold Standard", "Verra"],
    "data_cutoff": "2026-08-14",
    "introduction": intro,
    "structured_summary_rows": ss_rows,
    "validations": validations,
    "notes": notes,
    "change_log": changelog,
    "credit_row_count": len(rows),
}
with open(f"{OUT}/tool_meta.json", "w") as f:
    json.dump(meta, f, indent=1, ensure_ascii=False)

# ---------- console summary ----------
from collections import Counter
print("credit rows:", len(rows))
print("total volume:", sum(r["Volume"] or 0 for r in rows))
for col in ["Programme", "Country", "Vintage", "OIMP First Transfer Definition",
            "Authorisation Year", "Issuance Year", "Authorised for NDC?",
            "Authorised for (O)IMP?", "Authorisation uploaded to the CARP"]:
    print(f"\n{col}: {dict(Counter(str(r[col]) for r in rows).most_common())}")
print("\nnull check:", {c: sum(1 for r in rows if r[c] in (None, '')) for c in cols})
